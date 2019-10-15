import os
from selenium import webdriver
from openpyxl import load_workbook
import pandas as pd


def create_dir(path):
    '''
    新建文件夹
    :param path: 文件夹路径
    :return:
    '''
    # 删除已有文件
    if os.path.exists(path):  # 如果文件存在
        print('文件夹已存在！')
    else:
        os.makedirs(path)


def get_file_name(file_dir):
    '''
    获取文件夹下所有文件名
    :param file_dir:文件路径
    :return:当前路径下所有非目录子文件
    '''
    for root, dirs, files in os.walk(file_dir):
        # return root#当前目录路径
        # return dirs#当前路径下所有子目录
        return files


def del_exits_file(filename):
    if os.path.exists(filename):  # 如果文件存在
        # 删除文件，可使用以下两种方法。
        os.remove(filename)  # 则删除
        # os.unlink(my_file)


def from_A1_writer_to_excel_by_openpyxl(ls, table_path, sheetname):
    '''
    写入excel表格，从'A1'单元格开始
    :param ls:数据列表，形式为 [[a,b,c],[d,f,g],[h,i,j],...]
    :param table_path: 表路径名称
    :param sheetname: 新建表的名称
    :return:
    '''
    wb = load_workbook(table_path)
    sht = wb.create_sheet(sheetname)
    for row in range(1, len(ls) + 1):
        for col in range(1, len(ls[0]) + 1):
            sht.cell(row=row, column=col).value = ls[row - 1][col - 1]
    wb.save(table_path)
    wb.close()


def writer_branch_table(table_path, sheetname):
    '''
    筛选专本，并写入excel表格
    :param table_path: 补考报名表路径
    :param sheetname: 需要提取的学历等级，“专”或“本”
    :return:
    '''
    df = pd.read_excel(io=table_path, dtype=str, sheet_name='汇总')
    # "^\d"以数字开始，"{1}"一位(因为学号可能会出现0或1)，"[0-5]"数字“0-5”（6之前的），".*?"后面随意
    if sheetname == '本':
        re = '^\d{7}[2].*?'
    elif sheetname == '专':
        re = '^\d{7}[4].*?'
    else:
        print('只能输入“专”或“本”！')
    # 根据条件筛选出学号列
    student_nums = df['学号'].str.contains(re, na=False)
    df_new = df[student_nums]
    title = df_new.columns.tolist()  # 表头
    ls = []
    ls.append(title)  # 标题行作为首列表加入
    for i in range(df_new.shape[0]):
        ls.append(df_new.iloc[i].tolist())  # 每行数据转换为列表加入ls
    # 写入excel
    from_A1_writer_to_excel_by_openpyxl(ls, table_path, sheetname)



